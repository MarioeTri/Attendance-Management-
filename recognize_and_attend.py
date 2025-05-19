import cv2
import os
import numpy as np
from PIL import Image
import tkinter as tk
from tkinter import messagebox, messagebox
from datetime import datetime
import csv
from tkinter import filedialog, messagebox
import pandas as pd
import logging
import json
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import hashlib
from concurrent.futures import ThreadPoolExecutor
import tkinter as tk
from tkinter import messagebox, ttk
import time
import threading
import pyttsx3
import concurrent.futures
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from keras_facenet import FaceNet
from mtcnn import MTCNN

embedder = FaceNet()
detector = MTCNN()
engine = pyttsx3.init()
notification_lock = threading.Lock()
engine.setProperty('rate', 150)

attendance_threshold = {"hour": 8}
if os.path.exists("attendance_config.json"):
        with open("attendance_config.json", "r") as f:
            try:
                attendance_threshold = json.load(f)
            except json.JSONDecodeError:
                pass 
password_file = "admin_password.json"
if os.path.exists(password_file):
    with open(password_file, "r") as f:
        try:
            admin_password = json.load(f).get("password", "admin123")
        except json.JSONDecodeError:
            admin_password = "admin123"
else:
    admin_password = "admin123"
dataset_dir = 'dataset1(FaceNet) Shift'
model_path = 'trainer.yml'
csv_file = 'attendance7.csv'
labels_file = 'labels1(FaceNet).txt'
log_file = 'app.log'
admin_file = 'admin_pass.txt'
os.makedirs(dataset_dir, exist_ok=True)
logging.basicConfig(filename=log_file, level=logging.INFO, format='%(asctime)s %(message)s')
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()
def get_today_entry(name):
    if not os.path.exists(csv_file):
        return None, []
    today = datetime.now().strftime('%Y-%m-%d')
    with open(csv_file, 'r', newline='') as f:
        reader = list(csv.reader(f))
    if not reader:
        return None, []
    header = reader[0]
    rows = reader[1:]
    for i, row in enumerate(rows):
        if row[0] == name and row[1].startswith(today):
            return i + 1, reader 
    return None, reader
def is_admin(username, password):
    if not os.path.exists(admin_file):
        return False
    hashed = hash_password(password)
    with open(admin_file, 'r') as f:
        for line in f:
            u, p = line.strip().split(',')
            if u == username and p == hashed:
                return True
    return False
def register_admin(username, password):
    with open(admin_file, 'a') as f:
        f.write(f"{username},{hash_password(password)}\n")
def speak_notification(message):
    with notification_lock:  
        engine.stop()  
        engine.say(message)  
        engine.runAndWait()
def get_today_entry(name):
    if not os.path.exists(csv_file):
        return None, []
    today = datetime.now().strftime('%Y-%m-%d')
    with open(csv_file, 'r', newline='') as f:
        reader = list(csv.reader(f))
    if not reader:
        return None, []
    header = reader[0]
    rows = reader[1:]

    for i, row in enumerate(rows):
        if row[0] == name and row[1].startswith(today):
            return i + 1, reader
    return None, reader
def mark_attendance(name, mode="auto"):
    now = datetime.now()
    time_str = now.strftime('%Y-%m-%d %H:%M:%S')
    shift = get_user_shift(name)
    current_shift = determine_shift(now)
    
    if shift != current_shift:
        warning_message = f"Anda salah shift! Terdaftar untuk shift {shift}, tetapi mencoba check-in pada shift {current_shift}."
        print(warning_message)
        threading.Thread(target=speak_notification, args=(warning_message,)).start()
        return

    shift_data = attendance_threshold.get(shift, None)
    if not shift_data:
        print(f"Shift {shift} tidak ditemukan dalam pengaturan!")
        return
    shift_hour = shift_data["hour"]
    shift_minute = shift_data["minute"]
    checkin_start_time = now.replace(hour=shift_hour - 1, minute=shift_minute, second=0, microsecond=0)
    checkin_end_time = now.replace(hour=shift_hour, minute=shift_minute, second=0, microsecond=0)
    late_checkin_end_time = checkin_end_time + timedelta(hours=4)
    if checkin_start_time <= now <= late_checkin_end_time:
        if now <= checkin_end_time + timedelta(minutes=5):
            status = "Tepat Waktu"
        else:
            status = "Terlambat"
        row_index, data = get_today_entry(name)
        if row_index is None:
            new_row = [name, time_str, status, '', shift]
            if not os.path.exists(csv_file):
                with open(csv_file, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Name', 'TimeCheckIn', 'StatusCheckIn', 'TimeCheckOut', 'Shift'])
            with open(csv_file, 'a', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(new_row)
            log_type = "Check-in"
        else:
            existing_row = data[row_index]
            check_in_time = datetime.strptime(existing_row[1], '%Y-%m-%d %H:%M:%S')
            if existing_row[3]:
                print(f"{name} sudah melakukan check-out hari ini.")
                return
            if now >= check_in_time + timedelta(minutes=1):
                data[row_index][3] = time_str
                with open(csv_file, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerows(data)
                log_type = "Check-out"
            else:
                print(f"{name} sudah check-in, belum waktunya check-out.")
                return
    else:
        print(f"Waktu check-in untuk shift {shift} sudah lewat (lebih dari 4 jam).")
        return
    notification_message = f"{name} {log_type} ({shift}) - Status: {status}"
    print(notification_message)
    threading.Thread(target=speak_notification, args=(notification_message,)).start()
def determine_shift(current_time):
    if (current_time.hour >= attendance_threshold["pagi"]["hour"] - 1 and current_time.hour < attendance_threshold["siang"]["hour"] - 1):
        return "pagi"
    elif (current_time.hour >= attendance_threshold["siang"]["hour"] - 1 and current_time.hour < attendance_threshold["malam"]["hour"] - 1):
        return "siang"
    else:
        return "malam"
def detect_faces_mtcnn(image):
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    equalized_image = cv2.equalizeHist(gray_image)
    denoised_image = cv2.GaussianBlur(equalized_image, (5, 5), 0)
    results = detector.detect_faces(image)
    detected_faces = []
    for result in results:
        if result['confidence'] > 0.9:  
            detected_faces.append(result)
    return detected_faces
def register_face(name, shift):
    person_dir = os.path.join(dataset_dir, name)
    if os.path.exists(person_dir):
        messagebox.showerror("Error", "Nama sudah terdaftar. Gunakan nama lain.")
        return
    os.makedirs(person_dir, exist_ok=True)
    with open(os.path.join(person_dir, "shift.txt"), "w") as f:
        f.write(shift)
    cam = cv2.VideoCapture(0)
    count = 0
    while True:
        ret, frame = cam.read()
        if not ret:
            break
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)
        for (x, y, w, h) in faces:
            count += 1
            face_img = gray[y:y+h, x:x+w]
            cv2.imwrite(f"{person_dir}/{name}_{count}.jpg", face_img)
            cv2.rectangle(frame, (x,y), (x+w, y+h), (255,0,0), 2)
            cv2.putText(frame, f"{count}/40", (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255,255,255), 2)
        cv2.imshow('Register - Press q to quit', frame)
        if count >= 40 or cv2.waitKey(1) & 0xFF == ord('q'):
            break
    cam.release()
    cv2.destroyAllWindows()
    messagebox.showinfo("Done", f"Captured {count} faces for {name}")
    def save_embeddings_for_user(name):
        person_dir = os.path.join(dataset_dir, name)
        embeddings = []
        for image_name in os.listdir(person_dir):
            if not image_name.endswith(".jpg"):
                continue
            img_path = os.path.join(person_dir, image_name)
            img = cv2.imread(img_path)
            emb = get_face_embedding(img)
            if emb is not None:
                embeddings.append(emb)
        np.save(os.path.join(person_dir, f'{name}_embedding.npy'), embeddings)
    save_embeddings_for_user(name)
def change_user_shift(name, new_shift):
    person_dir = os.path.join(dataset_dir, name)
    shift_file = os.path.join(person_dir, "shift.txt")
    if not os.path.exists(person_dir):
        messagebox.showerror("Error", f"User '{name}' tidak ditemukan.")
        return
    with open(shift_file, "w") as f:
        f.write(new_shift)
    messagebox.showinfo("Sukses", f"Shift untuk {name} berhasil diubah ke {new_shift}")

def get_user_shift(name):
    shift_file = os.path.join(dataset_dir, name, "shift.txt")
    if os.path.exists(shift_file):
        with open(shift_file, "r") as f:
            return f.read().strip()
    return "pagi" 

def train_model():
    faces = []
    labels = []
    label_dict = {}
    current_id = 0
    valid_extensions = ('.jpg', '.jpeg', '.png', '.bmp') 
    for person_name in os.listdir(dataset_dir):
        person_dir = os.path.join(dataset_dir, person_name)
        if not os.path.isdir(person_dir):
            continue
        label_dict[current_id] = person_name
        for image_name in os.listdir(person_dir):
            if not image_name.lower().endswith(valid_extensions):
                continue  
            image_path = os.path.join(person_dir, image_name)
            try:
                img = Image.open(image_path).convert('L')
                img_np = np.array(img, 'uint8')
                faces.append(img_np)
                labels.append(current_id)
            except Exception as e:
                print(f"Error loading image {image_path}: {e}")
        current_id += 1
    with open(labels_file, 'w') as f:
        for id, name in label_dict.items():
            f.write(f"{id},{name}\n")
def load_labels():
    label_dict = {}
    if os.path.exists(labels_file):
        with open(labels_file, 'r') as f:
            for line in f:
                id, name = line.strip().split(',')
                label_dict[int(id)] = name
    return label_dict
def get_face_embedding(image):
    results = detect_faces_mtcnn(image)
    if results:
        x, y, w, h = results[0]['box']
        face = image[y:y+h, x:x+w]
        face = cv2.resize(face, (160, 160))
        face_rgb = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)
        emb = embedder.embeddings([face_rgb])[0]
        return emb
    return None
def get_batch_face_embeddings(faces):
    embeddings = []
    faces_rgb = [cv2.cvtColor(face, cv2.COLOR_BGR2RGB) for face in faces]
    embeddings = embedder.embeddings(faces_rgb)  
    return embeddings
def recognize_faces_facenet_with_threads():
    if not os.path.exists(dataset_dir):
        messagebox.showerror("Error", "Belum ada data.")
        return
    cam = cv2.VideoCapture(0)
    known_embeddings = {}
    for person_name in os.listdir(dataset_dir):
        emb_path = os.path.join(dataset_dir, person_name, f'{person_name}_embedding.npy')
        if os.path.exists(emb_path):
            known_embeddings[person_name] = np.load(emb_path)
    recognized = []
    with ThreadPoolExecutor(max_workers=8) as executor: 
        while True:
            ret, frame = cam.read()
            if not ret:
                break
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            futures = []
            future = executor.submit(detect_faces_mtcnn, rgb)
            faces_results = future.result()
            faces = []
            for res in faces_results:
                x, y, w, h = res['box']
                face = frame[y:y+h, x:x+w]
                faces.append(face)
            if faces:
                future_embeddings = executor.submit(get_batch_face_embeddings, faces)
                embeddings = future_embeddings.result()
                for emb, res in zip(embeddings, faces_results):
                    emb = emb / np.linalg.norm(emb)
                    x, y, w, h = res['box']
                    min_dist = float('inf')
                    identity = "Unknown"
                    for name, known_embs in known_embeddings.items():
                        for e in known_embs:
                            dist = np.linalg.norm(emb - e)
                            if dist < min_dist:
                                min_dist = dist
                                identity = name
                    if min_dist < 0.6:  
                        if identity not in recognized:
                            mark_attendance(identity, mode="auto")
                            recognized.append(identity)
                        color = (0, 255, 0)
                        label = f"{identity} ({min_dist:.2f})"
                    else:
                        color = (0, 0, 255)
                        label = "Unknown"
                    cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
                    cv2.putText(frame, label, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2)
            cv2.imshow("Face Recognition - Press 'q' to Quit", frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
    cam.release()
    cv2.destroyAllWindows()
def delete_user(name):
    person_dir = os.path.join(dataset_dir, name)
    if os.path.exists(person_dir):
        for file in os.listdir(person_dir):
            os.remove(os.path.join(person_dir, file))
        os.rmdir(person_dir)
        messagebox.showinfo("Delete", f"Data untuk {name} telah dihapus.")
    else:
        messagebox.showerror("Error", f"Data untuk {name} tidak ditemukan.")
def export_to_excel(root):
    def choose_save_location():
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                df = pd.read_csv(csv_file)
                if df.empty:
                    messagebox.showinfo("Info", "Data absensi masih kosong.")
                    return
                
                start_date = start_date_entry.get().strip()
                end_date = end_date_entry.get().strip()

                try:
                    start_date = datetime.strptime(start_date, "%Y-%m-%d")
                    end_date = datetime.strptime(end_date, "%Y-%m-%d")
                except ValueError:
                    messagebox.showerror("Error", "Format tanggal tidak valid. Gunakan format YYYY-MM-DD.")
                    return

                df['TimeCheckIn'] = pd.to_datetime(df['TimeCheckIn'])
                filtered_df = df[(df['TimeCheckIn'] >= start_date) & (df['TimeCheckIn'] <= end_date)]

                if filtered_df.empty:
                    messagebox.showinfo("Info", "Tidak ada data untuk tanggal yang dipilih.")
                    return

                filtered_df['Weekday'] = filtered_df['TimeCheckIn'].dt.day_name()
                filtered_df['Date'] = filtered_df['TimeCheckIn'].dt.date

                wb = openpyxl.Workbook()

                grouped = filtered_df.groupby(['Date', 'Shift'])

                for (date, shift), group in grouped:
                    sheet_name = f"{date} {shift.capitalize()}"
                    ws = wb.create_sheet(title=sheet_name)

                    ws.merge_cells('A1:F1')
                    title_cell = ws['A1']
                    title_cell.value = "Data Absensi PT MATTEL"
                    title_cell.font = Font(bold=True, size=14)
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")

                    ws.append([])

                    ws['A2'] = 'Tanggal'
                    ws['A2'].font = Font(bold=True)
                    ws['B2'] = str(date)

                    ws['A3'] = 'Shift'
                    ws['A3'].font = Font(bold=True)
                    ws['B3'] = shift.capitalize()

                    ws['E2'] = 'Kalkulasi Tepat Waktu'
                    ws['F2'] = f"{group[group['StatusCheckIn'].str.lower() != 'terlambat'].shape[0]}"
                    ws['E3'] = 'Kalkulasi Terlambat'
                    ws['F3'] = f"{group[group['StatusCheckIn'].str.lower() == 'terlambat'].shape[0]}"

                    terlambat_count = group[group['StatusCheckIn'].str.lower() == 'terlambat'].shape[0]
                    tepatwaktu_count = group[group['StatusCheckIn'].str.lower() != 'terlambat'].shape[0]

                    ws.append([])

                    header = ['Nama', 'TimeCheckIn', 'StatusCheckIn', 'TimeCheckOut', 'Shift', 'Weekday']
                    ws.append(header)

                    header_font = Font(bold=True, color="FFFFFF", size=12)
                    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    for col_num, header_cell in enumerate(ws[5], 1):
                        header_cell.font = header_font
                        header_cell.fill = header_fill
                        header_cell.alignment = Alignment(horizontal="center", vertical="center")

                    for _, row in group.iterrows():
                        is_late = row['StatusCheckIn'].lower() == 'terlambat'
                        row_data = [row['Name'], row['TimeCheckIn'], row['StatusCheckIn'], row['TimeCheckOut'], row['Shift'], row['Weekday']]
                        ws.append(row_data)

                        if is_late:
                            for col in range(1, 7):
                                cell = ws.cell(row=ws.max_row, column=col)
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=len(header)):
                        for cell in row:
                            cell.border = Border(left=Side(border_style="thin", color="000000"),
                                                 right=Side(border_style="thin", color="000000"),
                                                 top=Side(border_style="thin", color="000000"),
                                                 bottom=Side(border_style="thin", color="000000"))

                    for col in range(1, len(header) + 1):
                        col_letter = openpyxl.utils.get_column_letter(col)
                        ws.column_dimensions[col_letter].width = 20
                    ws.freeze_panes = 'A6'

                del wb['Sheet']

                shift_counts = filtered_df['Shift'].value_counts()

                chart_ws = wb.create_sheet(title="Perbandingan Shift")
                chart_ws['A1'] = "Shift"
                chart_ws['B1'] = "Jumlah Kehadiran"
                row = 2
                shifts = ["pagi", "siang", "malam"]
                for shift in shifts:
                    chart_ws.cell(row=row, column=1, value=shift)
                    chart_ws.cell(row=row, column=2, value=shift_counts.get(shift, 0))
                    row += 1

                chart = BarChart()
                data = Reference(chart_ws, min_col=2, min_row=1, max_col=2, max_row=row - 1)
                categories = Reference(chart_ws, min_col=1, min_row=2, max_row=row - 1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                chart.title = "Perbandingan Kehadiran Pagi, Siang, Malam"
                chart.x_axis.title = "Shift"
                chart.y_axis.title = "Jumlah Kehadiran"
                chart.style = 10
                chart.width = 25
                chart.height = 15
                chart.series[0].graphicalProperties.line.solidFill = "4F81BD"
                chart_ws.add_chart(chart, "D5")

                wb.save(file_path)

                messagebox.showinfo("Sukses", f"Data berhasil diekspor ke {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Terjadi kesalahan: {e}")

    date_window = tb.Toplevel(root)
    date_window.title("Pilih Rentang Tanggal")
    date_window.geometry("400x200")

    tb.Label(date_window, text="Masukkan tanggal mulai (YYYY-MM-DD):", bootstyle="info").pack(pady=5)
    start_date_entry = tb.Entry(date_window, width=30)
    start_date_entry.pack(pady=5)

    tb.Label(date_window, text="Masukkan tanggal akhir (YYYY-MM-DD):", bootstyle="info").pack(pady=5)
    end_date_entry = tb.Entry(date_window, width=30)
    end_date_entry.pack(pady=5)

    tb.Button(date_window, text="Pilih Lokasi dan Ekspor", command=choose_save_location, bootstyle="success").pack(pady=10)
def show_today_attendance():
    if not os.path.exists(csv_file):
        messagebox.showerror("Error", "File absensi tidak ditemukan.")
        return

    today = datetime.now().strftime('%Y-%m-%d')

    with open(csv_file, 'r') as f:
        lines = f.readlines()

    if not lines or len(lines) == 1:
        messagebox.showinfo("Hari Ini", "Belum ada data absensi.")
        return

    headers = lines[0].strip().split(",")
    data = [line.strip().split(",") for line in lines[1:] if today in line]

    if not data:
        messagebox.showinfo("Hari Ini", "Belum ada yang hadir hari ini.")
        return

    pagi_data = [row for row in data if row[4] == "pagi"]
    siang_data = [row for row in data if row[4] == "siang"]
    malam_data = [row for row in data if row[4] == "malam"]

    status_checkin_index = headers.index("StatusCheckIn")  

    top = tk.Toplevel()
    top.title(f"Data Absensi - {today}")
    top.geometry("700x450")
    top.configure(bg="#f0f0f0")

    label = ttk.Label(top, text=f"Daftar Kehadiran Hari Ini ({today})", font=('Segoe UI', 14, 'bold'))
    label.pack(pady=10)

    notebook = ttk.Notebook(top)
    notebook.pack(fill="both", expand=True, padx=15, pady=10)

    def create_tab(title, data):
        frame = ttk.Frame(notebook)
        notebook.add(frame, text=title)
        
        tree = ttk.Treeview(frame, columns=headers, show="headings")
        tree.pack(side="left", fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        
        for header in headers:
            tree.heading(header, text=header, command=lambda _h=header: sort_column(tree, _h, False))
            tree.column(header, anchor='center', width=120, stretch=True)
        
        for i, row in enumerate(data):
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'

            if row[status_checkin_index].lower() == "terlambat":
                tag = 'late'  

            tree.insert('', 'end', values=row, tags=(tag,))

        style = ttk.Style()
        style.configure("Treeview.Heading", font=('Segoe UI', 10, 'bold'))
        style.configure("Treeview", font=('Segoe UI', 10), rowheight=25)
        style.map("Treeview", background=[("selected", "#d1e7dd")])
        tree.tag_configure('evenrow', background='#ffffff')
        tree.tag_configure('oddrow', background='#f5f5f5')
        tree.tag_configure('late', background='#ffcccc')  

    create_tab("Pagi", pagi_data)
    create_tab("Siang", siang_data)
    create_tab("Malam", malam_data)

    def choose_shift_and_export():
        export_dialog = tk.Toplevel(top)
        export_dialog.title("Pilih Shift untuk Ekspor")
        export_dialog.geometry("300x200")
        
        selected_shift = tk.StringVar(value="semua")

        label = ttk.Label(export_dialog, text="Pilih shift yang ingin diekspor:", font=('Segoe UI', 12))
        label.pack(pady=10)

        shift_options = [("Semua Shift", "semua"), ("Pagi", "pagi"), ("Siang", "siang"), ("Malam", "malam")]
        for text, value in shift_options:
            radio_button = ttk.Radiobutton(export_dialog, text=text, variable=selected_shift, value=value)
            radio_button.pack(anchor="w", padx=20, pady=5)

        def export_selected_shift():
            shift = selected_shift.get()
            export_dialog.destroy()
            
            if shift == "pagi":
                data_to_export = pagi_data
            elif shift == "siang":
                data_to_export = siang_data
            elif shift == "malam":
                data_to_export = malam_data
            else: 
                data_to_export = pagi_data + siang_data + malam_data  

            if not data_to_export:
                messagebox.showinfo("Tidak Ada Data", f"Tidak ada data untuk shift {shift}.")
                return
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Simpan File Excel",
                initialfile=f"Absensi_{shift}_{today}.xlsx"
            )
            
            if not file_path:
                return

            from openpyxl import Workbook
            wb = Workbook()

            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

            def add_shift_sheet(wb, shift_name, data):
                ws = wb.create_sheet(title=shift_name.capitalize())

                ws.merge_cells('A1:E1')
                ws['A1'] = "Data Absensi PT MATTEL"
                ws['A1'].font = Font(bold=True, size=14)
                ws['A1'].alignment = Alignment(horizontal="center")

                ws.merge_cells('A2')
                ws['A2'] = f"Hari: "
                ws['A2'].font = Font(bold=True)
                ws.merge_cells('B2')
                ws['B2'] = f"{today}"
                ws['B2'].font = Font(bold=True)
                
                terlambat_count = sum(1 for row in data if row[status_checkin_index].lower() == "terlambat")
                ws.merge_cells('D2')
                ws['D2'] = f"Terlambat: "
                ws['D2'].font = Font(bold=True)
                ws.merge_cells('E2')
                ws['E2'] = f"{terlambat_count}"
                ws['E2'].font = Font(bold=True)

                ws.merge_cells('A3')
                ws['A3'] = f"Shift:"
                ws['A3'].font = Font(bold=True)
                ws.merge_cells('B3')
                ws['B3'] = f"{shift_name.capitalize()}"
                ws['B3'].font = Font(bold=True)

                tepatwaktu_count = len(data) - terlambat_count
                ws.merge_cells('D3')
                ws['D3'] = f"Tepat Waktu: "
                ws['D3'].font = Font(bold=True)
                ws.merge_cells('E3')
                ws['E3'] = f"{tepatwaktu_count}"
                ws['E3'].font = Font(bold=True)

                ws.append([])
                ws.append(headers)

                for row in data:
                    ws.append(row)

                for i, row in enumerate(data, 6):  
                    if row[status_checkin_index].lower() == "terlambat":
                        for col in range(1, len(headers) + 1):
                            ws.cell(row=i, column=col).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

                for col in range(1, len(headers) + 1):
                    col_letter = openpyxl.utils.get_column_letter(col)
                    ws.column_dimensions[col_letter].width = 20  

                ws.freeze_panes = 'A6'

            if shift == "semua":
                add_shift_sheet(wb, "Pagi", pagi_data)
                add_shift_sheet(wb, "Siang", siang_data)
                add_shift_sheet(wb, "Malam", malam_data)
            else:
                add_shift_sheet(wb, shift, data_to_export)
            
            wb.save(file_path)
            messagebox.showinfo("Berhasil", f"Data berhasil diekspor ke:\n{file_path}")

        export_button = ttk.Button(export_dialog, text="Ekspor", command=export_selected_shift)
        export_button.pack(pady=20)

    export_btn = ttk.Button(top, text="Export ke Excel", command=choose_shift_and_export)
    export_btn.pack(pady=5)

def sort_column(tree, col, reverse):
    l = [(tree.set(k, col), k) for k in tree.get_children('')]
    try:
        l.sort(key=lambda t: float(t[0]) if t[0].replace('.', '', 1).isdigit() else t[0], reverse=reverse)
    except:
        l.sort(reverse=reverse)

    for index, (val, k) in enumerate(l):
        tree.move(k, '', index)
    tree.heading(col, command=lambda: sort_column(tree, col, not reverse))
def run_gui():
    def open_shift_change_window():
        window = tk.Toplevel()
        window.title("Ganti Shift")
        window.geometry("400x350")

        tk.Label(window, text="Pilih Shift Awal:").pack(pady=5)
        initial_shift_var = tk.StringVar()
        shift_combo = ttk.Combobox(window, textvariable=initial_shift_var, values=["pagi", "siang", "malam"], state="readonly")
        shift_combo.pack(pady=5)

        tk.Label(window, text="Pilih Nama:").pack(pady=5)
        name_frame = tk.Frame(window)
        name_frame.pack(pady=5, fill="both", expand=True)

        name_listbox = tk.Listbox(name_frame, selectmode="multiple", height=8, exportselection=0)
        name_listbox.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(name_frame, orient="vertical", command=name_listbox.yview)
        scrollbar.pack(side="right", fill="y")
        name_listbox.config(yscrollcommand=scrollbar.set)

        select_all_var = tk.BooleanVar()
        select_all_check = ttk.Checkbutton(
            window, 
            text="Pilih Semua", 
            variable=select_all_var, 
            command=lambda: select_all_names(name_listbox, select_all_var)
        )
        select_all_check.pack(pady=5)

        tk.Label(window, text="Pilih Shift Baru:").pack(pady=5)
        new_shift_var = tk.StringVar()
        new_shift_combo = ttk.Combobox(window, textvariable=new_shift_var, values=["pagi", "siang", "malam"], state="readonly")
        new_shift_combo.pack(pady=5)

        def update_name_list(*args):
            selected_shift = initial_shift_var.get()
            name_listbox.delete(0, tk.END)
            select_all_var.set(False)
            if selected_shift:
                names = [
                    name for name in os.listdir(dataset_dir)
                    if os.path.isdir(os.path.join(dataset_dir, name)) and
                    get_user_shift(name).lower() == selected_shift.lower()
                ]
                for name in names:
                    name_listbox.insert(tk.END, name)
                select_all_check.config(state="normal" if names else "disabled")
            else:
                select_all_check.config(state="disabled")

        def select_all_names(listbox, var):
            if var.get():
                listbox.select_set(0, tk.END)
            else:
                listbox.selection_clear(0, tk.END)

        def apply_change():
            selected_shift = initial_shift_var.get()
            selected_names = [name_listbox.get(i) for i in name_listbox.curselection()]
            new_shift = new_shift_var.get()
            
            if not selected_shift:
                messagebox.showerror("Error", "Pilih shift awal terlebih dahulu.")
                return
            if not selected_names:
                messagebox.showerror("Error", "Pilih setidaknya satu nama pengguna.")
                return
            if not new_shift:
                messagebox.showerror("Error", "Pilih shift baru.")
                return
                
            for name in selected_names:
                change_user_shift(name, new_shift)
            messagebox.showinfo("Sukses", f"Shift untuk {len(selected_names)} pengguna berhasil diubah ke {new_shift}")
            window.destroy()

        initial_shift_var.trace('w', update_name_list)
        tk.Button(window, text="Ubah Shift", command=apply_change).pack(pady=10)
    def open_register_window():
        def on_register():
            name = name_entry.get().strip()
            shift = shift_var.get()
            if not name or shift == "":
                messagebox.showerror("Error", "Nama dan Shift wajib diisi!")
                return
            register_face(name, shift)
            window.destroy()

        window = tk.Toplevel()
        window.title("Registrasi Wajah")
        window.geometry("300x200")

        tk.Label(window, text="Nama:").pack(pady=5)
        name_entry = tk.Entry(window)
        name_entry.pack(pady=5)

        tk.Label(window, text="Shift:").pack(pady=5)
        shift_var = tk.StringVar()
        shift_combo = ttk.Combobox(window, textvariable=shift_var, values=["pagi", "siang", "malam"], state="readonly")
        shift_combo.pack(pady=5)

        tk.Button(window, text="Daftar", command=on_register).pack(pady=10)
    def on_train():
        train_model()
        messagebox.showinfo("Train", "Model training selesai!")
    def on_recognize():
        recognize_faces_facenet_with_threads()
    def on_delete():
        name = name_entry_admin.get().strip()
        if name:
            delete_user(name)
        else:
            messagebox.showerror("Error", "Masukkan nama untuk dihapus!")
    def on_export():
        export_to_excel()
    def on_today():
        show_today_attendance()
    def set_attendance_time():
        global attendance_threshold
        def save_time():
            try:
                pagi_hour = int(pagi_hour_combobox.get())
                pagi_minute = int(pagi_minute_combobox.get())
                siang_hour = int(siang_hour_combobox.get())
                siang_minute = int(siang_minute_combobox.get())
                malam_hour = int(malam_hour_combobox.get())
                malam_minute = int(malam_minute_combobox.get())
                if 0 <= pagi_hour <= 23 and 0 <= pagi_minute <= 59 and \
                0 <= siang_hour <= 23 and 0 <= siang_minute <= 59 and \
                0 <= malam_hour <= 23 and 0 <= malam_minute <= 59:
                    attendance_threshold["pagi"] = {"hour": pagi_hour, "minute": pagi_minute}
                    attendance_threshold["siang"] = {"hour": siang_hour, "minute": siang_minute}
                    attendance_threshold["malam"] = {"hour": malam_hour, "minute": malam_minute}
                    max_shift_hour = max(malam_hour, siang_hour, pagi_hour)
                    max_shift_minute = max(malam_minute, siang_minute, pagi_minute)
                    attendance_threshold["shift_start_time"] = {
                        "hour": (max_shift_hour - 1) % 24,
                        "minute": max_shift_minute
                    }
                    with open("attendance_config.json", "w") as f:
                        json.dump(attendance_threshold, f)
                    messagebox.showinfo("Berhasil", f"Pengaturan shift telah diset.")
                    config_window.destroy()
                else:
                    messagebox.showerror("Error", "Jam, menit tidak valid.")
            except ValueError:
                messagebox.showerror("Error", "Pilih jam, menit yang valid.")
        config_window = tb.Toplevel(root)
        config_window.title("Pengaturan Shift dan Jam Masuk")
        config_window.geometry("400x350")
        frame_main = tb.Frame(config_window, bootstyle="light")
        frame_main.pack(padx=20, pady=20, fill="both", expand=True)
        tb.Label(frame_main, text="Pengaturan Shift", font=("Segoe UI", 16, "bold"), bootstyle="info").pack(pady=10)
        tb.Label(frame_main, text="Pilih jam masuk shift pagi:", font=("Segoe UI", 12), bootstyle="secondary").pack(pady=5)
        pagi_hours = [str(i).zfill(2) for i in range(24)]
        pagi_hour_combobox = ttk.Combobox(frame_main, values=pagi_hours, state="readonly", font=("Segoe UI", 14))
        pagi_hour_combobox.set(str(attendance_threshold.get("pagi", {}).get("hour", 8)).zfill(2))
        pagi_hour_combobox.pack(pady=5, ipadx=5, ipady=5)
        pagi_minutes = [str(i).zfill(2) for i in range(60)]
        pagi_minute_combobox = ttk.Combobox(frame_main, values=pagi_minutes, state="readonly", font=("Segoe UI", 14))
        pagi_minute_combobox.set(str(attendance_threshold.get("pagi", {}).get("minute", 0)).zfill(2))
        pagi_minute_combobox.pack(pady=5, ipadx=5, ipady=5)
        tb.Label(frame_main, text="Pilih jam masuk shift siang:", font=("Segoe UI", 12), bootstyle="secondary").pack(pady=5)
        siang_hour_combobox = ttk.Combobox(frame_main, values=pagi_hours, state="readonly", font=("Segoe UI", 14))
        siang_hour_combobox.set(str(attendance_threshold.get("siang", {}).get("hour", 14)).zfill(2))
        siang_hour_combobox.pack(pady=5, ipadx=5, ipady=5)
        siang_minute_combobox = ttk.Combobox(frame_main, values=pagi_minutes, state="readonly", font=("Segoe UI", 14))
        siang_minute_combobox.set(str(attendance_threshold.get("siang", {}).get("minute", 0)).zfill(2))
        siang_minute_combobox.pack(pady=5, ipadx=5, ipady=5)
        tb.Label(frame_main, text="Pilih jam masuk shift malam:", font=("Segoe UI", 12), bootstyle="secondary").pack(pady=5)
        malam_hour_combobox = ttk.Combobox(frame_main, values=pagi_hours, state="readonly", font=("Segoe UI", 14))
        malam_hour_combobox.set(str(attendance_threshold.get("malam", {}).get("hour", 22)).zfill(2))
        malam_hour_combobox.pack(pady=5, ipadx=5, ipady=5)
        malam_minute_combobox = ttk.Combobox(frame_main, values=pagi_minutes, state="readonly", font=("Segoe UI", 14))
        malam_minute_combobox.set(str(attendance_threshold.get("malam", {}).get("minute", 0)).zfill(2))
        malam_minute_combobox.pack(pady=5, ipadx=5, ipady=5)
        tb.Button(frame_main, text="Simpan", command=save_time, bootstyle="success", width=20).pack(pady=10)
        tb.Button(frame_main, text="Batal", command=config_window.destroy, bootstyle="danger", width=20).pack(pady=5)
    def show_salary_window():
        salary_window = tb.Toplevel()
        salary_window.title("Salary Calculation")
        salary_window.geometry("700x500")
        salary_window.resizable(False, False)
        input_frame = tb.LabelFrame(salary_window, text="Input Data Karyawan", bootstyle="primary")
        input_frame.pack(fill="x", padx=20, pady=10)
        tb.Label(input_frame, text="Nama Karyawan:", bootstyle="info").grid(row=0, column=0, sticky="w", padx=10, pady=5)
        name_entry = tb.Entry(input_frame, width=30)
        name_entry.grid(row=0, column=1, padx=10, pady=5)
        tb.Label(input_frame, text="Gaji per Jam (Rp):", bootstyle="info").grid(row=1, column=0, sticky="w", padx=10, pady=5)
        wage_entry = tb.Entry(input_frame, width=30)
        wage_entry.grid(row=1, column=1, padx=10, pady=5)
        tb.Label(input_frame, text="Tanggal Mulai (YYYY-MM-DD):", bootstyle="info").grid(row=2, column=0, sticky="w", padx=10, pady=5)
        start_entry = tb.Entry(input_frame, width=30)
        start_entry.grid(row=2, column=1, padx=10, pady=5)
        tb.Label(input_frame, text="Tanggal Akhir (YYYY-MM-DD):", bootstyle="info").grid(row=3, column=0, sticky="w", padx=10, pady=5)
        end_entry = tb.Entry(input_frame, width=30)
        end_entry.grid(row=3, column=1, padx=10, pady=5)
        result_frame = tb.Frame(salary_window)
        result_frame.pack(pady=10, fill="both", expand=True)
        def calculate_salary():
            name = name_entry.get().strip()
            wage = wage_entry.get().strip()
            start = start_entry.get().strip()
            end = end_entry.get().strip()
            if not name or not wage or not start or not end:
                messagebox.showwarning("Warning", "Isi semua data terlebih dahulu.")
                return
            try:
                hourly_rate = float(wage)
                start_date = datetime.strptime(start, "%Y-%m-%d")
                end_date = datetime.strptime(end, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("Error", "Masukkan data yang valid.")
                return
            try:
                df = pd.read_csv(csv_file)
                df.columns = df.columns.str.strip()
                df = df[df['Name'].str.strip().str.lower() == name.lower()]
                if df.empty:
                    messagebox.showerror("Error", f"Nama {name} tidak ditemukan.")
                    return
                df['TimeCheckIn'] = pd.to_datetime(df['TimeCheckIn'])
                df['TimeCheckOut'] = pd.to_datetime(df['TimeCheckOut'], errors='coerce')
                df = df[(df['TimeCheckIn'] >= start_date) & (df['TimeCheckIn'] <= end_date)]
                df['Duration'] = (df['TimeCheckOut'] - df['TimeCheckIn']).dt.total_seconds() / 3600
                df['Duration'] = df['Duration'].fillna(0).clip(lower=0)
                df['Gaji'] = df['Duration'] * hourly_rate
                total_salary = df['Gaji'].sum()
                for widget in result_frame.winfo_children():
                    widget.destroy()
                tb.Label(result_frame, text=f"💼 Total Gaji: Rp {total_salary:,.2f}",
                        bootstyle="success", font=("Helvetica", 13, "bold")).pack(pady=10)
                columns = ['Name', 'TimeCheckIn', 'TimeCheckOut', 'StatusCheckIn', 'Duration', 'Gaji']
                tree_frame = tb.Frame(result_frame)
                tree_frame.pack(fill="both", expand=True)
                tree_scroll = ttk.Scrollbar(tree_frame)
                tree_scroll.pack(side="right", fill="y")
                tree = ttk.Treeview(tree_frame, columns=columns, show="headings", yscrollcommand=tree_scroll.set)
                tree_scroll.config(command=tree.yview)
                for col in columns:
                    tree.heading(col, text=col)
                    tree.column(col, width=110, anchor="center")
                for _, row in df.iterrows():
                    tree.insert("", "end", values=[
                        row['Name'],
                        row['TimeCheckIn'].strftime("%Y-%m-%d %H:%M:%S") if pd.notnull(row['TimeCheckIn']) else '',
                        row['TimeCheckOut'].strftime("%Y-%m-%d %H:%M:%S") if pd.notnull(row['TimeCheckOut']) else '',
                        row.get('StatusCheckIn', ''),
                        f"{row['Duration']:.2f}",
                        f"Rp {row['Gaji']:.2f}"
                    ])
                tree.pack(fill="both", expand=True)
            except Exception as e:
                messagebox.showerror("Error", f"Terjadi kesalahan saat menghitung gaji:\n{e}")
        tb.Button(salary_window, text="🔍 Calculate", command=calculate_salary,
                bootstyle="primary outline").pack(pady=10)
    def show_all_attendance():
        if not os.path.exists(csv_file):
            messagebox.showerror("Error", "File absensi tidak ditemukan.")
            return

        df = pd.read_csv(csv_file)
        if df.empty:
            messagebox.showinfo("Absensi", "Data absensi masih kosong.")
            return

        df['Date'] = pd.to_datetime(df['TimeCheckIn']).dt.date
        df['Weekday'] = pd.to_datetime(df['TimeCheckIn']).dt.day_name()

        pagi_df = df[df['Shift'].str.lower() == 'pagi']
        siang_df = df[df['Shift'].str.lower() == 'siang']
        malam_df = df[df['Shift'].str.lower() == 'malam']

        top = tk.Toplevel(root)
        top.title("Semua Data Absensi")
        top.geometry("1000x550")
        top.configure(bg="#f0f0f0")

        title = ttk.Label(top, text="Data Absensi Keseluruhan", font=('Segoe UI', 14, 'bold'))
        title.pack(pady=10)

        notebook = ttk.Notebook(top)
        notebook.pack(fill="both", expand=True, padx=15, pady=10)

        def create_tab(title, dataframe):
            frame = ttk.Frame(notebook)
            notebook.add(frame, text=title)

            tree = ttk.Treeview(frame, columns=list(dataframe.columns), show="headings")
            tree.pack(side="left", fill="both", expand=True)

            scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side="right", fill="y")

            for col in dataframe.columns:
                tree.heading(col, text=col, anchor='center')
                tree.column(col, anchor='center', width=120, stretch=True)

            for i, row in dataframe.iterrows():
                tag = 'evenrow' if i % 2 == 0 else 'oddrow'

                if row['StatusCheckIn'].lower() == 'terlambat':
                    tag = 'late'

                tree.insert('', 'end', values=list(row), tags=(tag,))

            style = ttk.Style()
            style.configure("Treeview.Heading", font=('Segoe UI', 10, 'bold'))
            style.configure("Treeview", font=('Segoe UI', 10), rowheight=25)
            style.map("Treeview", background=[("selected", "#d1e7dd")])

            tree.tag_configure('evenrow', background='#ffffff')
            tree.tag_configure('oddrow', background='#f5f5f5')
            tree.tag_configure('late', background='red', foreground='white')

        create_tab("Pagi", pagi_df)
        create_tab("Siang", siang_df)
        create_tab("Malam", malam_df)

        export_btn = ttk.Button(top, text="Ekspor ke Excel", command=lambda: export_to_excel(df))
        export_btn.pack(pady=10)

        search_frame = ttk.Frame(top)
        search_frame.pack(pady=5)

        search_label = ttk.Label(search_frame, text="Cari Nama atau Tanggal:")
        search_label.pack(side="left", padx=(0, 5))

        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
        search_entry.pack(side="left")

        def filter_data():
            keyword = search_var.get().lower().strip()
            filtered_pagi_df = pagi_df[pagi_df.apply(lambda row: row.astype(str).str.lower().str.contains(keyword).any(), axis=1)]
            filtered_siang_df = siang_df[siang_df.apply(lambda row: row.astype(str).str.lower().str.contains(keyword).any(), axis=1)]
            filtered_malam_df = malam_df[malam_df.apply(lambda row: row.astype(str).str.lower().str.contains(keyword).any(), axis=1)]

            for widget in notebook.winfo_children():
                widget.destroy()

            create_tab("Pagi", filtered_pagi_df)
            create_tab("Siang", filtered_siang_df)
            create_tab("Malam", filtered_malam_df)

        search_btn = ttk.Button(search_frame, text="Cari", command=filter_data)
        search_btn.pack(side="left", padx=5)

        reset_btn = ttk.Button(search_frame, text="Reset", command=lambda: filter_data())
        reset_btn.pack(side="left")
        
        
    def switch_to_admin():
        def verify_password():
            entered_password = password_entry.get()
            if entered_password == admin_password:
                messagebox.showinfo("Akses Diberikan", "Berhasil masuk sebagai admin.")
                notebook.add(admin_tab, text="Admin")
                notebook.select(admin_tab)
                pw_window.destroy()
            else:
                messagebox.showerror("Akses Ditolak", "Password salah!")
        pw_window = tb.Toplevel(root)
        pw_window.title("Verifikasi Admin")
        pw_window.geometry("300x150")
        tb.Label(pw_window, text="Masukkan Password Admin:", bootstyle="info").pack(pady=10)
        password_entry = tb.Entry(pw_window, show="*", width=25)
        password_entry.pack(pady=5)
        tb.Button(pw_window, text="Masuk", command=verify_password, bootstyle="primary").pack(pady=10)
        password_entry.focus()
    def change_admin_password():
        def verify_old_password():
            if old_pass_entry.get() == admin_password:
                old_pass_window.destroy()
                enter_new_password()
            else:
                messagebox.showerror("Error", "Password lama salah!")
        def enter_new_password():
            def save_new_password():
                new_pass = new_pass_entry.get()
                if len(new_pass) < 4:
                    messagebox.showerror("Error", "Password terlalu pendek (min. 4 karakter)")
                    return
                with open(password_file, "w") as f:
                    json.dump({"password": new_pass}, f)
                messagebox.showinfo("Berhasil", "Password admin berhasil diganti!")
                new_pass_window.destroy()
                global admin_password
                admin_password = new_pass
            new_pass_window = tb.Toplevel(root)
            new_pass_window.title("Password Baru")
            new_pass_window.geometry("300x150")
            tb.Label(new_pass_window, text="Masukkan Password Baru:", bootstyle="info").pack(pady=10)
            new_pass_entry = tb.Entry(new_pass_window, show="*", width=25)
            new_pass_entry.pack(pady=5)
            tb.Button(new_pass_window, text="Simpan", command=save_new_password, bootstyle="success").pack(pady=10)
        old_pass_window = tb.Toplevel(root)
        old_pass_window.title("Verifikasi Password Lama")
        old_pass_window.geometry("300x150")
        tb.Label(old_pass_window, text="Masukkan Password Lama:", bootstyle="warning").pack(pady=10)
        old_pass_entry = tb.Entry(old_pass_window, show="*", width=25)
        old_pass_entry.pack(pady=5)
        tb.Button(old_pass_window, text="Lanjut", command=verify_old_password, bootstyle="primary").pack(pady=10)
    def reset_attendance():
        if os.path.exists(csv_file):
            if messagebox.askyesno("Konfirmasi", "Yakin ingin menghapus seluruh data absensi?"):
                os.remove(csv_file)
                messagebox.showinfo("Reset", "Data absensi telah dihapus.")
        else:
            messagebox.showerror("Error", "File absensi tidak ditemukan.")
    def show_graph_window():
        if not os.path.exists(csv_file):
            messagebox.showerror("Error", "Belum ada data absensi.")
            return
        df = pd.read_csv(csv_file)
        if df.empty:
            messagebox.showinfo("Info", "Data absensi masih kosong.")
            return
        if 'StatusCheckIn' not in df.columns:
            messagebox.showerror("Error", "Kolom 'StatusCheckIn' tidak ditemukan.")
            return
        def plot_graph(name=None):
            ax.clear()
            if name:
                person_data = df[df["Name"].str.lower() == name.lower()]
                if person_data.empty:
                    messagebox.showinfo("Info", f"Tidak ada data untuk {name}")
                    return
                count_tepat = (person_data["StatusCheckIn"] == "Tepat Waktu").sum()
                count_terlambat = (person_data["StatusCheckIn"] == "Terlambat").sum()
                title = f"Grafik Kehadiran - {name.title()}"
            else:
                count_tepat = (df["StatusCheckIn"] == "Tepat Waktu").sum()
                count_terlambat = (df["StatusCheckIn"] == "Terlambat").sum()
                title = "Grafik Kehadiran (Semua)"
            bars = ax.bar(
                ["Tepat Waktu", "Terlambat"],
                [count_tepat, count_terlambat],
                color=["#4CAF50", "#F44336"],
                edgecolor='black'
            )
            for bar in bars:
                height = bar.get_height()
                ax.annotate(f'{height}',
                            xy=(bar.get_x() + bar.get_width() / 2, height),
                            xytext=(0, 5),
                            textcoords="offset points",
                            ha='center', va='bottom', fontsize=10, fontweight='bold')

            ax.set_title(title, fontsize=14, fontweight='bold')
            ax.set_ylabel("Jumlah", fontsize=12)
            ax.set_xlabel("Status", fontsize=12)
            ax.grid(axis='y', linestyle='--', alpha=0.7)
            canvas.draw()
        def on_combo_select(event):
            selected = combo_name.get()
            if selected == "(Semua)":
                plot_graph()
            else:
                plot_graph(selected)
        def on_search():
            search_name = entry_search.get().strip()
            if not search_name:
                messagebox.showerror("Error", "Masukkan nama untuk dicari.")
                return
            plot_graph(search_name)
        graph_win = tb.Toplevel(root)
        graph_win.title("Grafik Kehadiran")
        graph_win.geometry("720x560")
        graph_win.configure(bg="#f8f9fa")
        frame_top = tb.Frame(graph_win)
        frame_top.pack(pady=10)
        tb.Label(frame_top, text="Pilih Nama:", bootstyle="secondary").grid(row=0, column=0, padx=5)
        names = sorted(df["Name"].unique().tolist())
        combo_name = ttk.Combobox(frame_top, values=["(Semua)"] + names, width=30)
        combo_name.set("(Semua)")
        combo_name.grid(row=0, column=1, padx=5)
        combo_name.bind("<<ComboboxSelected>>", on_combo_select)
        tb.Label(frame_top, text="Atau cari nama:", bootstyle="secondary").grid(row=1, column=0, padx=5, pady=5)
        entry_search = tb.Entry(frame_top, width=33)
        entry_search.grid(row=1, column=1, padx=5, pady=5)
        tb.Button(frame_top, text="Cari", command=on_search, bootstyle="info").grid(row=1, column=2, padx=5)
        fig, ax = plt.subplots(figsize=(6, 4))
        fig.tight_layout(pad=4.0)
        canvas = FigureCanvasTkAgg(fig, master=graph_win)
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        plot_graph()
    root = tb.Window(themename="flatly")
    root.title("Face Attendance (Extended)")
    root.geometry("480x740")
    notebook = tb.Notebook(root)
    notebook.pack(fill='both', expand=True, padx=10, pady=10)
    user_tab = tb.Frame(notebook)
    notebook.add(user_tab, text="User")
    tb.Label(user_tab, text="Face Recognition Attendance", font=("Arial", 16, "bold"), bootstyle="primary").pack(pady=20)
    tb.Button(user_tab, text="Mulai Absensi", command=on_recognize, bootstyle="success", width=25).pack(pady=5)
    tb.Button(user_tab, text="Pindah ke Admin", command=switch_to_admin, bootstyle="secondary", width=25).pack(pady=15)
    admin_tab = tb.Frame(notebook)
    tb.Label(admin_tab, text="Menu Admin", font=("Arial", 16, "bold"), bootstyle="dark").pack(pady=20)
    tb.Label(admin_tab, text="Nama untuk dihapus:", bootstyle="info").pack()
    name_entry_admin = tb.Entry(admin_tab, width=30)    
    name_entry_admin.pack(pady=5)
    tb.Button(admin_tab, text="Train Model", command=on_train, bootstyle="success", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Hapus Data Wajah", command=on_delete, bootstyle="danger", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Lihat Kehadiran Hari Ini", command=on_today, bootstyle="secondary", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Ganti Password Admin", command=change_admin_password, bootstyle="warning", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Set Jam Masuk", command=set_attendance_time, bootstyle="warning", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Lihat Semua Absensi", command=show_all_attendance, width=25).pack(pady=5)
    tb.Button(admin_tab, text="Reset Semua Absensi", command=reset_attendance, bootstyle="danger", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Register", command=open_register_window, bootstyle="danger", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Change Shift", command=open_shift_change_window, bootstyle="danger", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Salary Calculation", command=show_salary_window, bootstyle="secondary", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Lihat Grafik Absensi", command=show_graph_window, bootstyle="info", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Keluar", command=root.quit, bootstyle="dark", width=25).pack(pady=15)
    root.mainloop()
if __name__ == "__main__":
    run_gui()
